"""
import gspread
import gspread_dataframe as gd

gc = gspread.service_account(filename='./credentials.json')

# Open a sheet from a spreadsheet in one go
sht1 = gc.open_by_url("https://docs.google.com/spreadsheets/d/1TT7DbGj6F6BN10PS0PkSLAXXLyX9i-ILlBEs70X-Lac/edit?usp=sharing")
worksheet = sht1.sheet1


df = gd.get_as_dataframe(worksheet, parse_dates=True)

# Extract the cell background color information
colors_D = df.iloc[1:21, 3].style.background_color.tolist()
colors_E = df.iloc[1:21, 4].style.background_color.tolist()
print('Colors of cells D2:D22:', colors_D)
print('Colors of cells E2:E22:', colors_E)


#print(sht1)
"""
import gspread
from gspread_dataframe import set_with_dataframe
import openpyxl 

def get_color(rgb):
    if rgb == "FF00FF00":
        return "Green"
    elif rgb == "FF999999":
        return "Grey"
    elif rgb == "FFFF9900":
        return "Orange"
    elif rgb == "FFFFFF00":
        return "Yellow"
    elif rgb == "FF0000FF":
        return "Blue"
    elif rgb == "FFFF0000":
        return "Red"

def get_severity(rgb):
    if rgb == "FF0000FF":
        return 0
    elif rgb == "FF999999":
        return 1
    elif rgb == "FF00FF00":
        return 2
    elif rgb == "FFFFFF00":
        return 3
    elif rgb == "FFFF9900":
        return 4
    elif rgb == "FFFF0000":
        return 5

# Use the credentials to authenticate with the Google API
gc = gspread.service_account(filename='./credentials.json')

# Open a sheet from a spreadsheet in one go
sht1 = gc.open_by_url("https://docs.google.com/spreadsheets/d/1TT7DbGj6F6BN10PS0PkSLAXXLyX9i-ILlBEs70X-Lac/edit?usp=sharing")

export_file = sht1.export(format= gspread.utils.ExportFormat.EXCEL)
f = open('temp.xlsx', 'wb')
f.write(export_file)
f.close()

# Load the Excel file into a Workbook object using openpyxl
workbook = openpyxl.reader.excel.load_workbook('temp.xlsx')

# Get the first sheet of the Workbook
worksheet = workbook.worksheets[0]

# Find the names of places
name_place = []

r = 2
cell = worksheet.cell(row=r, column=1)
value = cell.value.strip()
name_place.append(value)
while value != None:
    r += 1
    cell = worksheet.cell(row=r, column=1)
    value = cell.value
    name_place.append(value)
    print(value)

colors_D = []
colors_E = []
sev_D = []
sev_E = []

# Loop through the cells to extract the cell background color information
for row in range(2, r):
    cell = worksheet.cell(row=row, column=4)
    colors_D.append(get_color(cell.fill.start_color.index))
    sev_D.append(get_severity(cell.fill.start_color.index))
    

for row in range(2, r):
    cell = worksheet.cell(row=row, column=5)
    colors_E.append(get_color(cell.fill.start_color.index))
    sev_E.append(get_severity(cell.fill.start_color.index))

result_malzeme = dict(zip(name_place, sev_D))
result_insan = dict(zip(name_place, sev_E))

print("Malzeme:\n", result_malzeme)
print("Insan:\n", result_insan)
print(len(result_malzeme))

