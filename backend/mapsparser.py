import csv
import random
import requests
import os
import django
import time
import re
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill

os.environ.setdefault("DJANGO_SETTINGS_MODULE", "afetgonullu.settings")
os.environ["DJANGO_SETTINGS_MODULE"] = "afetgonullu.settings"
django.setup()

from authentication.models import User
from locations.models import Location

# Store the list of created random variables and make sure there is no duplicate
random_variables = []

not_valid = []

df = pd.read_excel('data.xlsx')

# Load the Excel workbook
wb = openpyxl.load_workbook('data.xlsx')

# Get the first worksheet
ws = wb.worksheets[0]

# Open users.txt file
users_file = open('users.txt', 'w')

for index, row in df.iterrows():
    county = row[0]
    name = row[1]
    maps_url = row[3]

    print(row)

    print(county, name, maps_url)
    try:
        r = requests.get(maps_url)
    except:
        not_valid.append(index)
        continue
    print(r.url)
    match = re.search("@(-?\d+\.\d+),(-?\d+\.\d+),", r.url)

    if match:
        latitude = match.group(1)
        longitude = match.group(2)
    else:
        not_valid.append(index)
        continue

    # Get the cell in the 5th column of the current row
    cell = ws.cell(row=index + 1, column=5)

    # Get the background color of the cell
    background_color = cell.fill.start_color.index

    if background_color == 'FFFF0000':
        severity = 0
    elif background_color == 'FF00FF00':
        severity = 5
    else:
        severity = None

    activity = ""
    help_message = ""
    
    activity_row = str(row[4])
    help_row = str(row[6])

    if activity_row and activity_row != 'NaN' and activity_row != 'nan' and activity_row != '':
        activity = activity_row
    else:
        activity = ""
    if help_row and help_row != 'NaN' and help_row != 'nan' and help_row != '':
        help_message = help_row
    else:
        help_message = ""

    # If location already exists, update the severity and help_message
    if Location.objects.filter(name=name).exists():
        location = Location.objects.get(name=name)
        # If 5th column's background color is red set severity to 0,
        # if it is green set severity to 5
        
        if len(activity) != 0:
            location.activity = activity
            
        if len(help_message) != 0:
            location.help_message = help_message

        if severity is not None:
            location.severity = severity
            
        location.save()

    else:
        if severity is None:
            severity = 0
        location = Location.objects.create(county=county, 
                                            name=name, 
                                            maps_url=maps_url, latitude=latitude, activity=activity, help_message=help_message, longitude=longitude, update_automatically=True, severity=severity)
        
        
        random_integer = random.randint(0, 99999)
        # Make sure the random integer is not already in the list
        while random_integer in random_variables:
            random_integer = random.randint(0, 99999)
        random_variables.append(random_integer)

        email = f"deprem{random_integer}@gonullu.com"
        password = "1234567"
        print(email, password)
        user = User.objects.create_user(email=email, password=password)
        user.locations.add(location)
        user.save()

        # Add the user to users.txt file with location info
        users_file.write(f"{email} {password}: İstanbul {county} {name}")

    # Wait for a second
    time.sleep(1)
    users_file.write("\n")
    # Save users file and close
    users_file.flush()

# Write not_valid to notvalid.txt
with open('notvalid.txt', 'w') as f:
    for invalid in not_valid:
        f.write(str(invalid) + "\n")